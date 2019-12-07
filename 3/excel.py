from openpyxl import load_workbook
import requests
from bs4 import BeautifulSoup
from pymongo import MongoClient           # pymongo를 임포트 하기(패키지 인스톨 먼저 해야겠죠?)
client = MongoClient('localhost', 27017)  # mongoDB는 27017 포트로 돌아갑니다.
db = client.dbsparta                      # 'dbsparta'라는 이름의 db를 만듭니다.


# 타겟 URL을 읽어서 HTML를 받아오고,
headers = {'User-Agent' : 'Mozilla/5.0 (Windows NT 10.0; Win64; x64)AppleWebKit/537.36 (KHTML, like Gecko) Chrome/73.0.3683.86 Safari/537.36'}
data = requests.get('https://movie.naver.com/movie/sdb/rank/rmovie.nhn?sel=pnt&date=20190909',headers=headers)

# HTML을 BeautifulSoup이라는 라이브러리를 활용해 검색하기 용이한 상태로 만듦
# soup이라는 변수에 "파싱 용이해진 html"이 담긴 상태가 됨
# 이제 코딩을 통해 필요한 부분을 추출하면 된다.
soup = BeautifulSoup(data.text, 'html.parser')
table = soup.select_one('table.list_ranking')
movies = table.select('tr')

title_list = []
rating_list = []
rank_list = []
work_book = load_workbook('prac01.xlsx')
work_sheet = work_book['prac']

num = 0
for movie in movies:
    title = movie.select_one('td.title > div > a')
    rating = movie.select_one('td.point')
    rank = movie.select_one('td.ac > img')
    if title is not None:
        title_list.append(title.text)
        rating_list.append(rating.text)
        rank_list.append(rank['alt'])
        # 데이터를 입력합니다.
        work_sheet.cell(row=num + 2, column=1, value=rank_list[num])
        work_sheet.cell(row=num + 2, column=2, value=title_list[num])
        work_sheet.cell(row=num + 2, column=3, value=rating_list[num])
        num += 1

# 수정한 엑셀파일을 저장합니다.
# 참고: 다른이름으로 저장할 수도 있습니다.
work_sheet.cell(row=1, column=3, value='평점')
work_book.save('prac01.xlsx')

# MongoDB에 insert 하기

# 'users'라는 collection에 {'name':'bobby','age':21}를 넣습니다.
db.users.insert_one({'name':'bobby','age':21})
db.users.insert_one({'name':'kay','age':27})
db.users.insert_one({'name':'john','age':30})
